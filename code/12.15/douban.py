import re
import os
import requests
import openpyxl
from lxml import etree

# 爬取某一种类书籍的类
class CrawlBook():

    def __init__(self):
        self.crawl_page = 1
        self.next_path = ['']
        self.cookies = {
            'bid': 'QA_2hSwb-ko',
            '_vwo_uuid_v2': 'D67C0550837B8B5BF936D92E03BEECC38|b3323bb40f957b5f6a7d5bf5e76e6623',
            '__yadk_uid': 'HTMckBnNGWUQziQEGblskfW4ja0NRhSD',
            'll': '118267',
            'gr_user_id': 'ca82dc66-e0d8-4621-9ff7-861abbb8ab3b',
            'viewed': '33440205_6811366',
            '__utmc': '30149280',
            'push_noty_num': '0',
            'push_doumail_num': '0',
            '__utmv': '30149280.21585',
            '__utmz': '30149280.1608532007.15.12.utmcsr=douban.com|utmccn=(referral)|utmcmd=referral|utmcct=/',
            '_pk_ref.100001.8cb4': '%5B%22%22%2C%22%22%2C1608547763%2C%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3D7vs5g6VYlXuRMMS21RV_1J0eaCE-9RccWYmGqSi8q8iklNneLZ6m93FaFVMCJp2f%26wd%3D%26eqid%3Dca18b2b70002c6d5000000065fe034c1%22%5D',
            '_pk_ses.100001.8cb4': '*',
            '__utma': '30149280.199516907.1603362898.1608535640.1608547765.17',
            'dbcl2': '215857353:JnwSTmKMCRI',
            'ck': 'feNo',
            '__utmt': '1',
            '_pk_id.100001.8cb4': '6c8ca0915de026b0.1608028909.11.1608548529.1608532836.',
            '__gads': 'ID=21a27379861f62ad:T=1603596132:S=ALNI_MbOnNKj8ELwI-6StgMYXWtRXmK4Ww',
            '__utmb': '30149280.19.9.1608548530117',
        }

        self.headers = {
            'Connection': 'keep-alive',
            'Cache-Control': 'max-age=0',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-User': '?1',
            'Sec-Fetch-Dest': 'document',
            'Referer': 'https://www.douban.com/doumail/',
            'Accept-Language': 'zh-CN,zh;q=0.9',
        }

    # 获取书籍页
    def __getBookInfoPage(self):
        self.__readFile()
        while len(self.next_path):
            # next_path中前7个字符串已经包含在base_link中，所以只截取后面的与之拼接
            url = self.base_link+"?"+self.next_path.pop(0).split('?')[-1]
            r = requests.get(url, headers=self.headers, cookies=self.cookies)
            r.encoding = 'utf-8'

            html = etree.HTML(r.text)

            self.__getNextPage(html)
            self.__refine(html)
        self.__save()
        
        print("{} 爬取完毕！".format(self.classify_name))

    # 获取一下页链接
    def __getNextPage(self, html):
        try:
            next_url = html.xpath('//span[@class="next"]/a/@href')[0].strip()
            self.next_path.append(next_url)
        except IndexError:
            return

    # 提炼书籍页信息
    def __refine(self, html):
        book_nodes = html.xpath('//li[@class="subject-item"]')
        if book_nodes == None:
            self.next_path.clear()
            return

        # 遍历每本书，提取书本信息信息
        for book_node in book_nodes:
            book_name = book_node.xpath('./div[@class="info"]/h2/a/text()')[0].strip()
            # 提取的是作者，译者，出版社，出版时间，价格的字符串
            book_info = book_node.xpath('./div[@class="info"]/div[@class="pub"]/text()')[0].strip()
            
            try:
                star = book_node.xpath('./div[@class="info"]/div[@class="star clearfix"]/span[@class="rating_nums"]/text()')[0].strip()
            except IndexError:
                star = "本书无星评"
            # 书的摘要
            try:
                abstract = book_node.xpath('./div[@class="info"]/p/text()')[0].strip()
            except IndexError:
                abstract = '本书无简述'
            
            info_ls = book_info.split('/')
            """
                长度为5，则为外国书籍，有译者
                长度为4，则为国内书籍，无译者
                长度小于4，则不保存（此类极少）
            """
            if len(info_ls) < 4:
                continue

            author = info_ls[0].strip()
            translator = ""
            if len(info_ls) == 5:
                translator = info_ls[1]
            pub = info_ls[-3].strip()
            pub_date = info_ls[-2].strip()
            price = info_ls[-1].strip()

            data = [book_name, author, star, translator, pub, pub_date, price, abstract]
            # 存储
            self.__store(data)
        
        # 每爬取一页，报告一页
        print("page: {}".format(self.crawl_page))
        self.crawl_page += 1

    # 读excel文件
    def __readFile(self):
        self.wb = openpyxl.load_workbook("./spider/book.xlsx")
        if self.classify_name not in self.wb.sheetnames:
            self.wb.create_sheet(self.classify_name)
        self.sheet = self.wb[self.classify_name]
    
    # 保存excel文件
    def __save(self):
        self.wb.save("./spider/book.xlsx")
    
    # 存储一本书的书籍信息
    def __store(self, data):  
        self.sheet.append(data)

    # 某类书籍 爬虫入口
    def crawl(self, classify_path, classify_name):
        self.base_link = classify_path
        self.classify_name = classify_name
        self.__getBookInfoPage()
        print("退出书籍爬取类...")

# book = CrawlBook()
# book.crawl('https://book.douban.com/tag/小说', '小说')

# 爬取所有分类的URL
class Douban():
    def __init__(self):
        self.cookies = {
            'bid': 'QA_2hSwb-ko',
            '_vwo_uuid_v2': 'D67C0550837B8B5BF936D92E03BEECC38|b3323bb40f957b5f6a7d5bf5e76e6623',
            '__yadk_uid': 'HTMckBnNGWUQziQEGblskfW4ja0NRhSD',
            'll': '118267',
            'gr_user_id': 'ca82dc66-e0d8-4621-9ff7-861abbb8ab3b',
            'viewed': '33440205_6811366',
            '__utmc': '30149280',
            'push_noty_num': '0',
            'push_doumail_num': '0',
            '__utmv': '30149280.21585',
            '__utmz': '30149280.1608532007.15.12.utmcsr=douban.com|utmccn=(referral)|utmcmd=referral|utmcct=/',
            '_pk_ref.100001.8cb4': '%5B%22%22%2C%22%22%2C1608547763%2C%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3D7vs5g6VYlXuRMMS21RV_1J0eaCE-9RccWYmGqSi8q8iklNneLZ6m93FaFVMCJp2f%26wd%3D%26eqid%3Dca18b2b70002c6d5000000065fe034c1%22%5D',
            '_pk_ses.100001.8cb4': '*',
            '__utma': '30149280.199516907.1603362898.1608535640.1608547765.17',
            'dbcl2': '215857353:JnwSTmKMCRI',
            'ck': 'feNo',
            '__utmt': '1',
            '_pk_id.100001.8cb4': '6c8ca0915de026b0.1608028909.11.1608548529.1608532836.',
            '__gads': 'ID=21a27379861f62ad:T=1603596132:S=ALNI_MbOnNKj8ELwI-6StgMYXWtRXmK4Ww',
            '__utmb': '30149280.19.9.1608548530117',
        }
        self.headers = {
            'Connection': 'keep-alive',
            'Cache-Control': 'max-age=0',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-User': '?1',
            'Sec-Fetch-Dest': 'document',
            'Referer': 'https://www.douban.com/doumail/',
            'Accept-Language': 'zh-CN,zh;q=0.9',
        }
        self.base_url = 'https://book.douban.com'
    
    # 获取所有分类的url
    def __getClassifyUrl(self, html):
        type_url_dict = {}
        type_nodes = html.xpath('//table[@class="tagCol"]/tbody/tr/td/a')
        for type_node in type_nodes:
            type_link = type_node.xpath('./@href')[0].strip()
            type_name = type_node.xpath('./text()')[0].strip()
            type_url_dict[type_name] = type_link
        
        self.__fetchClassifyBooks(type_url_dict)

    # 通过CrawlBook类，按类爬取所有书籍信息
    def __fetchClassifyBooks(self, type_url_dict):
        
        for name, link in type_url_dict.items():
            if name in ["小说", "外国文学", "文学"]:
                continue
            url = self.base_url+link
            bookSpider = CrawlBook()
            bookSpider.crawl(url, name)
            print("Douban 类 爬取 {} 结束。。。".format(name))

    # 分类爬虫入口
    def go(self, url):
        r = requests.get(url=url, headers=self.headers, cookies=self.cookies)
        r.encoding = 'utf-8'
        html = etree.HTML(r.text)
        self.__getClassifyUrl(html)

douban = Douban()
douban.go("https://book.douban.com/tag/?view=type&icn=index-sorttags-all")
